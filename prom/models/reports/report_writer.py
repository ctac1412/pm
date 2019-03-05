# -*- coding: utf-8 -*-
###############################################################################
#    License, author and contributors information in:                         #
#    __manifest__.py file at the root folder of this module.                  #
###############################################################################
from odoo import fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import date, datetime
from lxml import etree
import logging
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment,PatternFill, NamedStyle, Font, Border, Side
from openpyxl.utils import get_column_letter
import io
import types
_logger = logging.getLogger("report_writer")

class report_writer(object):
    def __init__(self):
        self.master_row =1 
        self.master_cell = 1
        self.max_cell = 1
        self.wb = Workbook()
        self.ws = self.wb.active
        self.rus_months = ["Unknown",
          u"Январь",
          u"Февраль",
          u"Март",
          u"Апрель",
          u"Май",
          u"Июнь",
          u"Июль",
          u"Август",
          u"Сентябрь",
          u"Октябрь",
          u"Ноябрь",
          u"Декабрь"]
        self.registrate_style()
        

    def registrate_style(self):
        grennCell = NamedStyle(name="grennCell")
        # grennCell.font = Font(bold=True, size=20)
        grennCell.fill = PatternFill(start_color='dae6c0', end_color='dae6c0', fill_type='solid')
        grennCell.alignment = Alignment(wrap_text=True,horizontal='center',vertical="top")
        
        self.wb.add_named_style(grennCell)


    def render_payment_labels(self,payment_quarters):
        row = self.master_row
        self.cell_in_row(u'Статья',value_style=self.label_style())

        for x in payment_quarters:
            for m in x["months"]:
                f_cell = self.master_cell
                # cell = max([self.master_cell])
                self.cell_in_row(self.rus_months[m['record']['month']],f_row = row,megred_cell=1)
                self.cell_in_row(u'План',f_cell = f_cell, f_row = row+1)
                self.cell_in_row(u'Факт',f_cell = f_cell + 1, f_row = row+1)

            f_cell = self.master_cell
            self.cell_in_row(str(x['quarter']) + " " +  u'квартал', f_row = row,megred_cell=1)
            self.cell_in_row(u'План',f_cell = f_cell, f_row = row+1)
            self.cell_in_row(u'Факт',f_cell = f_cell + 1, f_row = row+1)

        f_cell = self.master_cell
        self.cell_in_row(u'Итого', f_row = row,megred_cell=1)
        self.cell_in_row(u'План',f_cell = f_cell, f_row = row+1)
        self.cell_in_row(u'Факт',f_cell = f_cell + 1, f_row = row+1)
        
        self.master_row +=2
        self.master_cell = 1

    def stream(self):
        return io.BytesIO(save_virtual_workbook(self.wb))

    def write_cell(self,row,column,value,alignment = False, megred_cell = 0,style_name=False):

        if type(value) == types.BooleanType and not value :
            value = ""
        
        try:
            value = datetime.strptime(value, '%Y-%m-%d').strftime("%d.%m.%Y")
            pass
        except :
            pass

        c = self.ws.cell(row=row, column= column,value=value)
        if alignment:
            c.alignment = alignment
        if megred_cell:
            self.ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column = column + megred_cell)
        if style_name:
            c.style= style_name 
        
        # _logger.info('--------------------- Cell type')
        # _logger.info(type(value))
        # _logger.info(type(value) == float)
        if type(value) == float:
            c.number_format = '#,##0.00'

        return c

    def label_style(self):
        return Alignment(wrap_text=True,horizontal='left',vertical="top")

    def value_style(self):
        return Alignment(wrap_text=True,horizontal='center',vertical="top")

    def value_green_style(self):
        return Alignment(wrap_text=True,horizontal='center',vertical="top")

    def label_value(self,label,value,f_row=False,f_cell=False,label_style=False,value_style=False,style_name=False):
        f_row,f_cell = (f_row or self.master_row,f_cell or self.master_cell)
        label_style,value_style = (label_style or self.label_style(),value_style or self.value_style())
        self.write_cell(f_row,f_cell,label,label_style,style_name=style_name)
        self.write_cell(f_row,f_cell+1,value,value_style,style_name=style_name)
        self.master_row = f_row+ 1
        self.max_cell = max(3,self.max_cell)

    def cell_in_row(self,value,f_row=False,f_cell=False,value_style=False,megred_cell = 0,style_name = False):
        f_row,f_cell = (f_row or self.master_row, f_cell or self.master_cell)
        value_style = value_style or self.value_style()

        self.write_cell(f_row,f_cell,value,value_style,megred_cell=megred_cell,style_name=style_name)
        self.master_cell = f_cell + 1 + megred_cell
        self.max_cell = max(self.max_cell, self.master_cell)

    def array_row(self,values=[],f_row=False,f_cell=False):
        f_row,f_cell = (f_row or self.master_row,f_cell or self.master_cell)
        arr_len_excel = len(values)+1
        for x in range(1,arr_len_excel):
            cell = values[x-1] 
            if type(cell) is not tuple:
                cell = (cell,False,False)
            self.write_cell(f_row,f_cell if x==1 else (f_cell + x - 1),cell[0],cell[1])
        self.master_row = f_row+ 1
        self.max_cell = max(arr_len_excel ,self.max_cell,)

    def get_obligations(self,passport,x,money):
        obligations = []
        arr=[]
        if money:
            arr = passport.obligation_ids.filtered(lambda p: len(p.obligation_type_money_id))
        else:
            arr = passport.obligation_ids.filtered(lambda p: len(p.obligation_type_id))

        # print "--------"
        # print passport.obligation_ids
        # print arr

        for obligation in arr:
            obligation_date =  fields.Datetime.from_string(obligation.obligation_date)
            price = (obligation.price * obligation.count) if (obligation_date and obligation_date.month == x.month and obligation_date.year == x.year) else 0
            price = self.env['prom.passport'].toRub(obligation.obligation_date,obligation.currency_id,price)

            obligations.append({
                'name':obligation.obligation_type_money_id.name if money else obligation.obligation_type_id.name,
                'price': price
            })

        return obligations
    