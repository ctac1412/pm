# -*- coding: utf-8 -*-
###############################################################################
#    License, author and contributors information in:                         #
#    __manifest__.py file at the root folder of this module.                  #
###############################################################################
from odoo import fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import date, datetime
from lxml import etree
import logging
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import io
import types
from report_writer import report_writer


_logger = logging.getLogger('cf_write')

class plan_zakupok(report_writer):

    def default_month(self,record):
        vals = {
                    'record':record,
                    'avance':{
                        'plan':0,
                        'fact':0
                    },
                    'fact':{
                        'plan':0,
                        'fact':0
                    },
                    'endpnr':{
                        'plan':0,
                        'fact':0
                    },
                    'message':{
                        'plan':0,
                        'fact':0
                    },
                    'obligations':[]
                }
        return vals

    def find_price(self,record,x,type):
        filter_result = record.passport_ids.filtered(lambda r: r.is_actual == True)
        passport = filter_result[0] if filter_result else False

        if not passport: return 0

        avance_date = fields.Datetime.from_string(passport.avance_date_of_payment)
        fact_date = fields.Datetime.from_string(passport.fact_date_of_payment)
        endpnr_date = fields.Datetime.from_string(passport.avance_date_of_payment)
        message_date = fields.Datetime.from_string(passport.message_date_of_payment)

        if type=='avance': return (passport.avance_summ_cur_rub_date_podpis or 0) if (avance_date and avance_date.month == x.month and avance_date.year == x.year) else 0
        if type=='fact': return (passport.fact_summ_cur_rub_date_podpis or 0) if (fact_date and fact_date.month == x.month and fact_date.year == x.year) else 0
        if type=='endpnr': return (passport.endpnr_summ_cur_rub_date_podpis or 0) if (endpnr_date and endpnr_date.month == x.month and endpnr_date.year == x.year) else 0
        if type=='message': return (passport.message_summ_cur_rub_date_podpis or 0) if (message_date and message_date.month == x.month and message_date.year == x.year) else 0

    def generate_payment_quarters(self,r,passport):

        payment_month_ids = self.root_passport.payment_month_ids


        payment_quarters = list(set([(x.year,x.payment_quarter)for x in payment_month_ids]))
        payment_quarters = sorted(payment_quarters,  key=lambda x: (x[0], x[1]))
        payment_quarters = [{'year':x[0],
                        'quarter':x[1]
                        } for x in payment_quarters]

        root_obligations=[]
        
        for q in payment_quarters:
            q['months'] = []

            res_x = payment_month_ids.filtered(lambda r: r.year == q['year'] and r.payment_quarter == q['quarter'])
            for x in res_x:
                podryads = []
                for pod in r.sub_project_ids:
                    pod_pas = pod.passport_ids.filtered(lambda r: r.is_actual == True)
                    if len(pod_pas) > 1: pod_pas = pod_pas[0]
                    if not len(pod_pas): continue

                    pod_pas_date_of_start = fields.Datetime.from_string(pod_pas.date_of_start)
                    pod_pas_avance_date = fields.Datetime.from_string(pod_pas.avance_date_of_payment)
                    pod_pas_fact_date = fields.Datetime.from_string(pod_pas.fact_date_of_payment)
                    pod_pas_endpnr_date = fields.Datetime.from_string(pod_pas.endpnr_date_of_payment)
                    pod_pas_message_date = fields.Datetime.from_string(pod_pas.message_date_of_payment)

                    
                    avance = (pod_pas.price_rub_actual or 0) if (
                        pod_pas_date_of_start and
                        pod_pas_date_of_start.month == x.month and 
                        pod_pas_date_of_start.year == x.year 
                        # message_date.month == pod_pas_message_date.month and 
                        # message_date.year == pod_pas_message_date.year
                        ) else 0
                    # avance = 
                    fact = 0
                    endpnr = 0
                    # message = (pod_pas.message_summ_cur_rub_date_podpis or 0) if (
                    #     pod_pas_message_date and
                    #     pod_pas_message_date.month == x.month and 
                    #     pod_pas_message_date.year == x.year 
                    #     # message_date.month == pod_pas_message_date.month and 
                    #     # message_date.year == pod_pas_message_date.year
                    #     ) else 0
                    message = 0

                    customer_company_id = pod.customer_company_id.name if pod.customer_company_id else ""
                    contractor_company_id = pod.contractor_company_id.name if pod.contractor_company_id else ""
                    contragents = customer_company_id + "/" + contractor_company_id


                    contract_number = pod_pas.contract_number or ""
                    specification_number = pod_pas.specification_number or ""
                    dogovor = contract_number + "/" + specification_number

                    nomenclatura = ', '.join([product.product_item_id.name for product in pod_pas.product_ids])
                    
                    
                    
                    podryads.append({
                        'name':pod.name,
                        'contragents':contragents,
                        'dogovor':dogovor,
                        'nomenclatura':nomenclatura,
                        'month_plan':sum([avance,fact,endpnr,message]),
                        'pod_pas_date_of_start': pod_pas_date_of_start,
                    })

                q['months'].append({
                    'record':x,
                    'podryads':podryads,
                    'podryads_summ':sum(pod['month_plan'] for pod in podryads)
                })

            q['months'] = sorted(q['months'],  key=lambda x: x['record'].month)

        return payment_quarters

    def render_block(self,r,sub=False,first=False):
        rus_months = ['Unknown',
          u'Январь',
          u'Февраль',
          u'Март',
          u'Апрель',
          u'Май',
          u'Июнь',
          u'Июль',
          u'Август',
          u'Сентябрь',
          u'Октябрь',
          u'Ноябрь',
          u'Декабрь']
        
        passport = r.passport_ids.filtered(lambda r: r.is_actual == True)
        if len(passport) > 1:
            raise UserError(u'Больше 1 актуального паспорта.')
        elif not len(passport):
            return 

        self.root_passport = passport

        

        payment_quarters = self.generate_payment_quarters(r,passport)


        self.label_value(u'План закупок',r.customer_company_id.name)
        self.label_value(u'С НДС, рублей на дату подписания',passport.price_rub_date_sign )
        date_of_pr_start = '' if not passport.date_of_pr_start else fields.Datetime.from_string(passport.date_of_pr_start).strftime('%d.%m.%Y')
        date_of_start = '' if not passport.date_of_start else fields.Datetime.from_string(passport.date_of_start).strftime('%d.%m.%Y')
        self.label_value(u'Период', '{}-{}'.format(date_of_pr_start,date_of_start))
        self.label_value(u'Ответственный менеджер',r.manager_user_id.name)

        row = self.master_row
        self.cell_in_row(u'Проект',value_style=self.label_style())
        self.cell_in_row(u'Поставщик/Исполнитель',value_style=self.label_style())
        self.cell_in_row(u'Договор/спецификация',value_style=self.label_style())
        self.cell_in_row(u'Номенклатура',value_style=self.label_style())

        for x in payment_quarters:
            for m in x['months']:
                f_cell = self.master_cell
                # cell = max([self.master_cell])
                self.cell_in_row(rus_months[m['record']['month']],f_row = row)
                self.cell_in_row(u'План',f_cell = f_cell, f_row = row+1)

            f_cell = self.master_cell
            self.cell_in_row(str(x['quarter']) + ' ' +  u'квартал', f_row = row)
            self.cell_in_row(u'План',f_cell = f_cell, f_row = row+1)

        f_cell = self.master_cell
        self.cell_in_row(u'Итого:',value_style=self.label_style())
        self.cell_in_row(u'Итого в валюте договора:',value_style=self.label_style(), f_cell = f_cell+1)
        self.cell_in_row(u'План',f_cell = f_cell, f_row = row+1)

        self.master_row += 2
        self.master_cell = 1



        self.master_cell = 5
        row = self.master_row
        max_row = self.master_row
        for x in payment_quarters:
            # Значения подряда к каждому месяцу
            # Значения подряда к каждому месяцу
            for m in x['months']:
                pod_count = 0
                cell = self.master_cell
                for pod in m['podryads']:
                    self.cell_in_row(pod['month_plan'],f_cell=cell, f_row = row+pod_count)
                    pod_count+=1
                max_row = max([max_row,pod_count])

            
            # Сумма подряда к каждому кварталу
            # Сумма подряда к каждому кварталу
            pod_count = 0
            cell = self.master_cell
            for pod in m['podryads']:
                value = sum(m['podryads'][pod_count]['month_plan'] for m in x['months'])
                self.cell_in_row(value,f_cell=cell, f_row = row+pod_count)
                pod_count+=1
            
        # Итоговая сумма подряда по всем кварталам
        # Итоговая сумма подряда по всем кварталам
        pod_count = 0
        cell = self.master_cell
        for pod in payment_quarters[0]["months"][0]['podryads']:
            value = sum(sum(m['podryads'][pod_count]['month_plan'] for m in x['months']) for x in payment_quarters)
            self.cell_in_row(value,f_cell=cell, f_row = row+pod_count)
            self.cell_in_row(passport.env['prom.passport'].fromRub(fields.Datetime.now(),passport.currency_id,value),f_cell=cell+1, f_row = row+pod_count)
            pod_count+=1


        # Названия подрядчиков
        # Названия подрядчиков
        pod_count = 0
        for pod in payment_quarters[0]["months"][0]['podryads']:
            self.cell_in_row(pod['name'],f_cell=1, f_row = row+pod_count)
            self.cell_in_row(pod['contragents'],f_cell=2, f_row = row+pod_count)
            self.cell_in_row(pod['dogovor'],f_cell=3, f_row = row+pod_count)
            self.cell_in_row(pod['nomenclatura'],f_cell=4, f_row = row+pod_count)
            pod_count+=1


        self.master_cell = 1
        self.master_row = row+pod_count

        self.cell_in_row(u'Итого за период:',value_style=self.label_style())
        self.master_cell = 5


        root_sum = 0
        for x in payment_quarters:
            for m in x['months']:
                self.cell_in_row(m['podryads_summ'])
            value = sum(m['podryads_summ'] for m in x['months'])
            root_sum += value
            self.cell_in_row(value)
        self.cell_in_row(root_sum)
        self.cell_in_row(passport.env['prom.passport'].fromRub(fields.Datetime.now(),passport.currency_id,root_sum))
              
            
  

            # self.compute_row(self.root_payment_quarters,u'Сальдо денежных потоков по операционной деятельности','root_saldo')

        self.ws.column_dimensions[get_column_letter(1)].width = 35
        self.ws.column_dimensions[get_column_letter(2)].width = 25
        self.ws.column_dimensions[get_column_letter(3)].width = 25
        self.ws.column_dimensions[get_column_letter(4)].width = 15


    def pod_rows(self,r,payment_quarters):
        for pod in r.sub_project_ids:
            self.compute_row(payment_quarters,u'Заказчик: {}-{}/Исполнитель: {}-{}'.format(pod.customer_company_id.name,pod.customer_company_id.vat,pod.contractor_company_id.name,pod.contractor_company_id.vat),'zero')
        if not r.sub_project_ids:
            self.label_value('','')

    def pod_compute_saldo(self,m,x):
        plateshi = sum([sum([
            m['avance']['plan'],
            m['fact']['plan'],
            m['endpnr']['plan'],
            m['message']['plan']
        ])for m in x['months']]) or 0
        obligations = sum([sum(obl['price'] or 0 for obl in m['obligations']) for m in x['months']]) or 0
        plateshi_podryad = sum([sum([
                            m['avance']['subpodryad']['plan'],
                            m['fact']['subpodryad']['plan'],
                            m['endpnr']['subpodryad']['plan'],
                            m['message']['subpodryad']['plan']
                        ])for m in x['months']]) or 0
        return plateshi - obligations - plateshi_podryad

    def pod_saldo(self,m):
        plateshi = sum([m['avance']['plan'],m['fact']['plan'],m['endpnr']['plan'],m['message']['plan']]) or 0
        obligations = sum(obl['price'] or 0 for obl in m['obligations']) or 0
        plateshi_podryad = sum([m['avance']['subpodryad']['plan'],m['fact']['subpodryad']['plan'],m['endpnr']['subpodryad']['plan'],m['message']['subpodryad']['plan']]) or 0
        podryad_saldo = plateshi - obligations - plateshi_podryad
        return podryad_saldo

    def compute_row(self,payment_quarters,label,root,ob = False,sub_m = False,sub_quar = False,sub_podryad=False,r = False):
        plan = 0
        
        row = self.master_row
        self.cell_in_row(label,value_style=self.label_style())
        for x in payment_quarters:
            for m in x['months']:
                if root == 'zero':
                    self.cell_in_row(0,f_row = row)
                    self.cell_in_row(0,f_row = row)
                elif root == 'root_saldo':
                    plateshi = sum([m['avance']['plan'],m['fact']['plan'],m['endpnr']['plan'],m['message']['plan']]) or 0
                    podryad_saldo = self.pod_saldo(m)
                    plateshi_podryad = sum([m['avance']['subpodryad']['plan'],m['fact']['subpodryad']['plan'],m['endpnr']['subpodryad']['plan'],m['message']['subpodryad']['plan']]) or 0
                    obligations = sum(obl['price'] or 0 for obl in m['obligations']) or 0


                    print '-------', m['record']['year'],m['record']['month'], plateshi , podryad_saldo ,plateshi_podryad,obligations
                    self.cell_in_row(plateshi + podryad_saldo  - plateshi_podryad - obligations,f_row = row)
                    self.cell_in_row(0,f_row = row)
                elif root == 'saldo':
                    podryad_saldo = self.pod_saldo(m)
                    self.cell_in_row(podryad_saldo,f_row = row)
                    self.cell_in_row(0,f_row = row)

                elif root == 'compute_summ':
                    if not sub_podryad:
                        self.cell_in_row(sum([m['avance']['plan'],m['fact']['plan'],m['endpnr']['plan'],m['message']['plan']]),f_row = row)
                        self.cell_in_row(sum([m['avance']['fact'],m['fact']['fact'],m['endpnr']['fact'],m['message']['fact']]),f_row = row)
                    else:
                        self.cell_in_row(sum([m['avance']['subpodryad']['plan'],m['fact']['subpodryad']['plan'],m['endpnr']['subpodryad']['plan'],m['message']['subpodryad']['plan']]),f_row = row)
                        self.cell_in_row(0,f_row = row)
                elif root == 'ob':
                    value = 0
                    if sub_m == m:
                        value = ob['price']
                    self.cell_in_row(value,f_row = row)
                    self.cell_in_row(0,f_row = row)
                elif root == 'compute_summ_obligations':
                    value = sum(obl['price'] or 0 for obl in m['obligations']) or 0
                    self.cell_in_row(value,f_row = row)
                    self.cell_in_row(0,f_row = row)
                else:
                    if not sub_podryad:
                        self.cell_in_row(m[root]['plan'],f_row = row)
                        self.cell_in_row(m[root]['fact'],f_row = row)
                    else:
                        self.cell_in_row(m[root]['subpodryad']['plan'],f_row = row)
                        self.cell_in_row(m[root]['subpodryad']['fact'],f_row = row)
                        
            if root == 'zero':
                self.cell_in_row(0,f_row = row)
                self.cell_in_row(0,f_row = row)
            elif root == 'root_saldo':
                plateshi = sum([sum([
                                        m['avance']['plan'],
                                        m['fact']['plan'],
                                        m['endpnr']['plan'],
                                        m['message']['plan']
                                    ])for m in x['months']]) or 0
                podryad_saldo = self.pod_compute_saldo(m,x)
                plateshi_podryad = sum([sum([
                                        m['avance']['subpodryad']['plan'],
                                        m['fact']['subpodryad']['plan'],
                                        m['endpnr']['subpodryad']['plan'],
                                        m['message']['subpodryad']['plan']
                                    ])for m in x['months']]) or 0
                obligations = sum([sum(obl['price'] or 0 for obl in m['obligations']) for m in x['months']]) or 0
                
                self.cell_in_row(plateshi + podryad_saldo  - plateshi_podryad - obligations,f_row = row)
                self.cell_in_row(0,f_row = row)
            elif root == 'saldo':
                    podryad_saldo = self.pod_compute_saldo(m,x)
                    self.cell_in_row(podryad_saldo,f_row = row)
                    self.cell_in_row(0,f_row = row)
            elif root == 'compute_summ':
                if not sub_podryad:
                    self.cell_in_row(sum([sum([
                                        m['avance']['plan'],
                                        m['fact']['plan'],
                                        m['endpnr']['plan'],
                                        m['message']['plan']
                                    ])for m in x['months']]),f_row = row)
                    self.cell_in_row(0,f_row = row)
                else:
                    self.cell_in_row(sum([sum([
                                        m['avance']['subpodryad']['plan'],
                                        m['fact']['subpodryad']['plan'],
                                        m['endpnr']['subpodryad']['plan'],
                                        m['message']['subpodryad']['plan']
                                    ])for m in x['months']]),f_row = row)
                    self.cell_in_row(0,f_row = row)
            elif root == 'compute_summ_obligations':
                self.cell_in_row(sum([sum(obl['price'] or 0 for obl in m['obligations']) for m in x['months']]),f_row = row)
                self.cell_in_row(0,f_row = row)
            elif root == 'ob':
                value = 0
                if sub_quar == x:
                    for mont_in_sub_quar in sub_quar['months']:
                        if ob in mont_in_sub_quar['obligations']:
                            value += ob['price']
                self.cell_in_row(value,f_row = row)
                self.cell_in_row(0,f_row = row)
            else:
                if not sub_podryad:
                    self.cell_in_row(sum(m[root]['plan'] for m in x['months']),f_row = row)
                    self.cell_in_row(sum(m[root]['fact'] for m in x['months']),f_row = row)
                else:
                    self.cell_in_row(sum(m[root]['subpodryad']['plan'] for m in x['months']),f_row = row)
                    self.cell_in_row(sum(m[root]['subpodryad']['fact'] for m in x['months']),f_row = row)

        self.master_row += 1
        self.master_cell = 1

    def render(self,r=False):
        if not r: return 
        self.podryad_saldo = 0
        self.root_r = r
        self.render_block(r,first=True)
        return self.ws
