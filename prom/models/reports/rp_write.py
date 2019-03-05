# -*- coding: utf-8 -*-
###############################################################################
#    License, author and contributors information in:                         #
#    __manifest__.py file at the root folder of this module.                  #
###############################################################################
from odoo import  fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import date, datetime
from lxml import etree
import logging
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import io
import types
from report_writer import report_writer
_logger = logging.getLogger("rp_write")

class rp_write(report_writer):
    def render_block(self,r):
        passport = r.passport_ids.filtered(lambda r: r.is_actual == True)
        if len(passport) > 1:
            raise UserError(u'Больше 1 актуального паспорта.')
            
        if r.kind_podryad == "main":
            self.label_value(u"БЮДЖЕТ ПРОЕКТА","")
        else:
            self.label_value(u"Продажа","",style_name = 'grennCell')

        self.label_value(u"Проект",r.name)
        self.label_value(u"Ответственный менеджер",r.manager_user_id.name)
        self.label_value(u"Исполнитель по проекту",r.contractor_company_id.name)
        self.label_value(u"Заказчик",r.customer_company_id.name)
        if passport: 
            self.label_value(u"Договор",passport.contract_number)
            self.label_value(u"Спецификация",passport.specification_number)
            self.label_value(u"Дата подписания",passport.date_of_signing)
            self.label_value(u"Предмет договора",", ".join([ (x.product_item_id.name if (x.product_item_id and x.product_item_id.name) else "") for x in passport.product_ids]))
            self.label_value(u"Валюта договора",passport.currency_id.name)
            self.label_value(u"Валюта расчетов по договору",passport.calculate_currency_id.name)
            self.label_value(u"Цена договора в валюте договора на дату подписания",passport.price_currency_id_date_sign)
            self.label_value(u"Цена договора в валюте РФ на дату подписания",passport.price_rub_date_sign)
            self.label_value(u"курс на дату заключения сделки",passport.currency_of_signing)
            self.label_value(u"Экспорт",passport.is_export)
            self.label_value(u"Ставка НДС, %НДС",r.contractor_company_id.get_nds(passport.date_of_signing) or 0)
            self.label_value(u"Сумма НДС в рублях на дату подписания",(passport.price_rub_date_sign or 0) - (passport.price_rub_date_sign_wonds or 0))
            self.label_value(u"Условие поставки",passport.term_of_delivery)
            self.label_value(u"Дата старта",passport.date_of_pr_start)
            self.label_value(u"Срок производства",passport.date_of_pr_production)
            self.label_value(u"Срок доставки",passport.date_of_delivery)
            self.label_value(u"Срок ПНР",passport.date_of_start)
            self.label_value(u"Срок поставки",passport.date_of_start)
            self.label_value(u"Гарантийный период",passport.date_of_warranty_end)
            self.array_row([
                (u"Условия оплаты:",self.label_style(),),
                (u"Процент",self.value_style(),),
                (u"Cумма в валюте договора",self.value_style(),),
                (u"Сумма в валюте РФ на дату подписания договора",self.value_style(),),
                (u"Ожидаемая дата платежа",self.value_style(),)])
            self.array_row([
                (u"Аванс",self.label_style(),),
                ("{}{}".format(passport.avance_contract_part_pr or "","%"),self.value_style(),),
                (passport.avance_summ_cur_contract,self.value_style(),),
                (passport.avance_summ_cur_rub_date_podpis,self.value_style(),),
                (passport.avance_payment_date,self.value_style(),)])
            self.array_row([
                (u"Оплата по уведомлению о готовности",self.label_style(),),
                ("{}{}".format(passport.message_contract_part_pr or "","%"),self.value_style(),),
                (passport.message_summ_cur_contract,self.value_style(),),
                (passport.message_summ_cur_rub_date_podpis,self.value_style(),),
                (passport.message_payment_date,self.value_style(),)])
            self.array_row([
                (u"Оплата по завершению ПНР",self.label_style(),),
                ("{}{}".format(passport.endpnr_contract_part_pr or "","%"),self.value_style(),),
                (passport.endpnr_summ_cur_contract,self.value_style(),),
                (passport.endpnr_summ_cur_rub_date_podpis,self.value_style(),),
                (passport.endpnr_payment_date,self.value_style(),)])
            self.array_row([
                (u"Оплата по факту поставки",self.label_style(),),
                ("{}{}".format(passport.fact_contract_part_pr or "","%"),self.value_style(),),
                (passport.fact_summ_cur_contract,self.value_style(),),
                (passport.fact_summ_cur_rub_date_podpis,self.value_style(),),
                (passport.fact_payment_date,self.value_style(),)])
            self.label_value(u"Вид расчетов","наличный" if passport.pay_kind == "cash" else "безналичный")

        self.ws.column_dimensions[get_column_letter(self.master_cell)].width = 52
        self.ws.column_dimensions[get_column_letter(self.master_cell+1)].width = 25
        self.ws.column_dimensions[get_column_letter(self.master_cell+2)].width = 25
        self.ws.column_dimensions[get_column_letter(self.master_cell+3)].width = 50
        self.ws.column_dimensions[get_column_letter(self.master_cell+4)].width = 25
        return True

    def render(self,r=False):
        if not r: return 
        self.render_block(r)
        for s in r.sub_project_ids:
            # self.master_row=1
            # self.master_cell += self.max_cell-1
            self.master_cell = 1
            # self.max_cell = 1
            self.render(s)
        return self.ws